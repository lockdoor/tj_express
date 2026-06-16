import os
import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break
from pathlib import Path
from tj_express.core.dbf import read_dbf
from tj_express.config import EXPRESS_PATH

THAI_MONTHS = {
    1: "มกราคม",
    2: "กุมภาพันธ์",
    3: "มีนาคม",
    4: "เมษายน",
    5: "พฤษภาคม",
    6: "มิถุนายน",
    7: "กรกฎาคม",
    8: "สิงหาคม",
    9: "กันยายน",
    10: "ตุลาคม",
    11: "พฤศจิกายน",
    12: "ธันวาคม"
}

def format_tax_id(tax_id: str) -> str:
    """Formats 13-digit tax ID into x-xxxx-xxxxx-xx-x format."""
    tax_id = str(tax_id).strip().replace("-", "")
    if len(tax_id) == 13:
        return f"{tax_id[0]}-{tax_id[1:5]}-{tax_id[5:10]}-{tax_id[10:12]}-{tax_id[12]}"
    return tax_id

def get_company_info(company_folder: str) -> dict:
    """
    Reads ISINFO.DBF to extract company name, tax ID and address.
    Falls back to defaults if not found.
    """
    dbf_dir = Path(EXPRESS_PATH) / company_folder
    isinfo_path = dbf_dir / "ISINFO.DBF"
    
    info = {
        "name": "บริษัท รินาระ โกลบอล จำกัด",
        "tax_id": "0105563149349",
        "address": "เลขที่ 616/12-13 ถนน สาธุประดิษฐ์ แขวงบางโพงพาง เขตยานนาวา กรุงเทพมหานคร 10120",
        "org_num": 0
    }
    
    if isinfo_path.exists():
        try:
            df = read_dbf(str(isinfo_path))
            if not df.empty:
                row = df.iloc[0]
                thinam = str(row.get("THINAM", info["name"])).strip()
                taxid = str(row.get("TAXID", info["tax_id"])).strip()
                addr1 = str(row.get("ADDR01", "")).strip()
                addr2 = str(row.get("ADDR02", "")).strip()
                orgnum = int(row.get("ORGNUM", 0))
                
                info["name"] = thinam
                info["tax_id"] = taxid
                info["address"] = f"{addr1} {addr2}".strip()
                info["org_num"] = orgnum
        except Exception:
            pass
            
    return info

def generate_tax_report(company_folder: str, year: int, month: int) -> dict:
    """
    Generates the sales tax report by reading and joining:
    - ARTRN.DBF (Sales transactions)
    - ARMAS.DBF (Customers master)
    - ARTRNRM.DBF (Remarks/notes for custom client names/TAXIDs)
    """
    dbf_dir = Path(EXPRESS_PATH) / company_folder
    
    artrn_path = dbf_dir / "ARTRN.DBF"
    armas_path = dbf_dir / "ARMAS.DBF"
    artrnrm_path = dbf_dir / "ARTRNRM.DBF"
    
    # Read core tables
    artrn = read_dbf(str(artrn_path))
    armas = read_dbf(str(armas_path))
    
    # Read remarks (notes) table if it exists
    if artrnrm_path.exists():
        artrnrm = read_dbf(str(artrnrm_path))
    else:
        artrnrm = pd.DataFrame(columns=['DOCNUM', 'SEQNUM', 'REMARK'])
        
    # Filter by date
    artrn['DOCDAT'] = pd.to_datetime(artrn['DOCDAT'])
    filtered = artrn[(artrn['DOCDAT'].dt.year == year) & (artrn['DOCDAT'].dt.month == month)].copy()
    
    # Filter allowed prefixes (IV = Invoice, HS = Cash Sale, SR = Sales Return)
    filtered = filtered[filtered['DOCNUM'].str[:2].isin(['IV', 'HS', 'SR'])]
    
    if filtered.empty:
        return {
            "records": [],
            "total_amount": 0.0,
            "total_vat": 0.0,
            "company_info": get_company_info(company_folder)
        }
        
    # Sort by DOCDAT
    filtered = filtered.sort_values('DOCDAT')
    
    # Merge with ARMAS
    merged = pd.merge(filtered, armas, on='CUSCOD', how='left', suffixes=('', '_armas'))
    
    # Extract custom customer names and TAXIDs from ARTRNRM
    # Notes are grouped by DOCNUM, with SEQNUM '@1' as name and '@2' as tax_id
    remarks_dict = {}
    if not artrnrm.empty:
        artrnrm['DOCNUM'] = artrnrm['DOCNUM'].astype(str).str.strip()
        artrnrm['SEQNUM'] = artrnrm['SEQNUM'].astype(str).str.strip()
        artrnrm['REMARK'] = artrnrm['REMARK'].astype(str).str.strip()
        
        for docnum, group in artrnrm.groupby('DOCNUM'):
            name_row = group[group['SEQNUM'] == '@1']
            tax_row = group[group['SEQNUM'] == '@2']
            remarks_dict[docnum] = {
                'name': name_row.iloc[0]['REMARK'] if not name_row.empty else None,
                'tax_id': tax_row.iloc[0]['REMARK'] if not tax_row.empty else None
            }

    records = []
    total_amount = 0.0
    total_vat = 0.0
    
    for idx, (_, row) in enumerate(merged.iterrows(), start=1):
        docnum = str(row['DOCNUM']).strip()
        prefix = docnum[:2]
        sign = -1 if prefix == 'SR' else 1
        
        # Determine client name and tax ID
        cuscod = str(row['CUSCOD']).strip() if pd.notna(row['CUSCOD']) else ""
        is_online = cuscod in ['LAZADA', 'SHOPEE', 'TIKTOK']
        
        customer_name = ""
        tax_id = ""
        
        if is_online and docnum in remarks_dict and remarks_dict[docnum]['name']:
            # Use custom note remark overrides
            customer_name = remarks_dict[docnum]['name']
            tax_id = remarks_dict[docnum]['tax_id'] or ""
        else:
            # Use ARMAS standard client info
            prenam = str(row['PRENAM']).strip() if pd.notna(row['PRENAM']) else ""
            cusnam = str(row['CUSNAM']).strip() if pd.notna(row['CUSNAM']) else ""
            customer_name = f"{prenam} {cusnam}".strip()
            if not customer_name:
                customer_name = cuscod or "เงินสด"
            tax_id = str(row['TAXID']).strip() if pd.notna(row['TAXID']) else ""
            
        tax_id = tax_id.replace("-", "").strip()
        
        # Determine branch/HQ status
        org_num = int(row['ORGNUM']) if pd.notna(row['ORGNUM']) else -1
        is_hq = False
        branch_code = ""
        
        # Only show branch/HQ info if we have a valid 13-digit Tax ID
        if tax_id and tax_id != "0000000000000" and len(tax_id) == 13:
            if org_num == 0:
                is_hq = True
            elif org_num > 0:
                branch_code = f"{org_num:05d}"
            else:
                is_hq = True  # Default to HQ if org_num not set
        
        # Calculate amount and VAT
        netval = float(row['NETVAL']) if pd.notna(row['NETVAL']) else (float(row['TOTAL']) - float(row['VATAMT']))
        vatamt = float(row['VATAMT']) if pd.notna(row['VATAMT']) else 0.0
        
        amount = round(netval * sign, 2)
        vat = round(vatamt * sign, 2)
        
        total_amount += amount
        total_vat += vat
        
        records.append({
            "no": idx,
            "date": row['DOCDAT'].strftime('%Y-%m-%d'),
            "docnum": docnum,
            "customer_name": customer_name,
            "tax_id": tax_id,
            "is_hq": is_hq,
            "branch_code": branch_code,
            "amount": amount,
            "vat": vat
        })
        
    return {
        "records": records,
        "total_amount": round(total_amount, 2),
        "total_vat": round(total_vat, 2),
        "company_info": get_company_info(company_folder)
    }

def export_tax_report_excel(company_folder: str, year: int, month: int, output_path: str):
    """
    Exports the sales tax report into a formatted Excel sheet matching the template.
    """
    data = generate_tax_report(company_folder, year, month)
    records = data["records"]
    company_info = data["company_info"]
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"ขาย{THAI_MONTHS[month][:3]}{str(year + 543)[2:]}"
    
    # Enable grid lines and set print titles (repeat rows 1 to 8 on each page)
    ws.views.sheetView[0].showGridLines = True
    ws.print_title_rows = '1:8'
    
    # Fonts & Styles
    font_family = "Angsana New"
    title_font_1 = Font(name=font_family, size=18, bold=True)
    title_font_2 = Font(name=font_family, size=16, bold=True)
    header_font = Font(name=font_family, size=14, bold=True)
    body_font = Font(name=font_family, size=14)
    summary_font = Font(name=font_family, size=14, bold=True)
    
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    
    gray_fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")
    
    thin_border_side = Side(style='thin', color='BFBFBF')
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    double_bottom_border = Border(top=thin_border_side, bottom=Side(style='double', color='000000'))
    
    title_row_height = 26.25

    # Write Title Section
    ws['A1'] = "รายงานภาษีขาย"
    ws['A1'].font = title_font_1
    ws.row_dimensions[1].height = title_row_height
    
    ws['A2'] = f"ประจำเดือน {THAI_MONTHS[month]} {year + 543}"
    ws['A2'].font = title_font_1
    ws.row_dimensions[2].height = title_row_height
    
    formatted_company_tax_id = format_tax_id(company_info["tax_id"])
    ws['A3'] = f"ชื่อผู้ประกอบกิจการ  :  {company_info['name']}    เลขประจำตัวผู้เสียภาษีอากร     {formatted_company_tax_id}"
    ws['A3'].font = title_font_1
    ws.row_dimensions[3].height = title_row_height
    
    is_hq_checked = "X" if company_info["org_num"] == 0 else " "
    is_br_checked = "X" if company_info["org_num"] > 0 else " "
    br_num_str = f"{company_info['org_num']:05d}" if company_info["org_num"] > 0 else "     "
    ws['A4'] = f"ชื่อสถานประกอบการ  :  {company_info['name']}     [ {is_hq_checked} ]   สำนักงานใหญ่    [ {is_br_checked} ]   สาขาที่ {br_num_str}"
    ws['A4'].font = title_font_2
    ws.row_dimensions[4].height = title_row_height
    
    ws['A5'] = f"ที่อยู่ :  {company_info['address']}"
    ws['A5'].font = title_font_2
    ws.row_dimensions[5].height = title_row_height

    # Merge cells for header
    ws.merge_cells("A1:I1")
    ws.merge_cells("A2:I2")
    ws.merge_cells("A3:I3")
    ws.merge_cells("A4:I4")
    ws.merge_cells("A5:I5")
    # Align center
    ws["A1"].alignment = center_align
    ws["A2"].alignment = center_align
    ws["A3"].alignment = center_align
    ws["A4"].alignment = center_align
    ws["A5"].alignment = center_align
    
    
    # Set Table Headers Values
    ws["A7"] = "ลำดับ"
    ws["A8"] = "ที่"
    ws["B7"] = "ใบกำกับภาษี"
    ws["B8"] = "วัน/เดือน/ปี"
    ws["C8"] = "เลขที่"
    ws["D7"] = "ชื่อบริษัท"
    ws["E7"] = "เลขประจำตัวผู้เสียภาษี "
    ws["F7"] = "สถานประกอบการ"
    ws["F8"] = "สนญ."
    ws["G8"] = "สาขา"
    ws["H7"] = "มูลค่าสินค้า"
    ws["H8"] = "หรือบริการ"
    ws["I7"] = "จำนวนเงินภาษี"
    ws["I8"] = "มูลค่าเพิ่ม"
    
    # Set Column Widths
    col_widths = {
        "A": 6, "B": 12, "C": 14, "D": 30, "E": 18, "F": 8, "G": 10, "H": 16, "I": 16
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # Style all cells in Row 7 and 8 BEFORE merging (avoids MergedCell read-only errors)
    for row in [7, 8]:
        for col_idx in range(1, 10):
            cell = ws.cell(row=row, column=col_idx)
            cell.font = header_font
            cell.alignment = center_align
            cell.fill = gray_fill
            cell.border = thin_border
            
    # Merge cells AFTER styling
    ws.merge_cells("B7:C7")
    ws.merge_cells("D7:D8")
    ws.merge_cells("E7:E8")
    ws.merge_cells("F7:G7")
    ws.merge_cells("H7:H8")
    ws.merge_cells("I7:I8")
            
    # Data Rows & Page Summary Logic
    current_row = 9
    page_start_row = 9
    items_on_page = 0
    page_sum_rows = []
    
    summary_border = Border(
        left=thin_border_side,
        right=thin_border_side,
        top=thin_border_side,
        bottom=Side(style='medium', color='000000') # bottom medium border as in Excel
    )
    
    for idx, rec in enumerate(records):
        ws.row_dimensions[current_row].height = 27
        ws.cell(row=current_row, column=1, value=rec["no"]).alignment = center_align
        
        # Format Date
        dt_val = datetime.datetime.strptime(rec["date"], "%Y-%m-%d")
        c_date = ws.cell(row=current_row, column=2, value=dt_val)
        c_date.number_format = 'yyyy-mm-dd'
        c_date.alignment = center_align
        
        ws.cell(row=current_row, column=3, value=rec["docnum"]).alignment = center_align
        ws.cell(row=current_row, column=4, value=rec["customer_name"]).alignment = left_align
        
        # Format Tax ID
        fmt_tid = format_tax_id(rec["tax_id"]) if rec["tax_id"] else ""
        fmt_tid = "" if fmt_tid == "0-0000-00000-00-0" else fmt_tid
        c_tax = ws.cell(row=current_row, column=5, value=fmt_tid)
        c_tax.alignment = center_align
        c_tax.number_format = '@'  # Force text format
        
        # Numeric values
        c_amt = ws.cell(row=current_row, column=8, value=rec["amount"])
        c_amt.number_format = '#,##0.00;(#,##0.00);"-";@'
        c_amt.alignment = right_align
        
        c_vat = ws.cell(row=current_row, column=9, value=f"=H{current_row}*7%")
        c_vat.number_format = '#,##0.00;(#,##0.00);"-";@'
        c_vat.alignment = right_align
        
        # Set fonts and borders for standard data row
        for col_idx in range(1, 10):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.font = body_font
            cell.border = thin_border
            
        current_row += 1
        items_on_page += 1
        
        # Check if we reached the maximum of 33 items per page
        if items_on_page == 33:
            # Write "รวมแต่ละหน้า" row
            ws.row_dimensions[current_row].height = 20
            ws.cell(row=current_row, column=2, value="รวมแต่ละหน้า").alignment = left_align
            
            c_pamt = ws.cell(row=current_row, column=8, value=f"=SUM(H{page_start_row}:H{current_row-1})")
            c_pamt.number_format = '#,##0.00;(#,##0.00);"-";@'
            c_pamt.alignment = right_align
            c_pamt.font = summary_font
            
            c_pvat = ws.cell(row=current_row, column=9, value=f"=SUM(I{page_start_row}:I{current_row-1})")
            c_pvat.number_format = '#,##0.00;(#,##0.00);"-";@'
            c_pvat.alignment = right_align
            c_pvat.font = summary_font
            
            for col_idx in range(1, 10):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.font = summary_font if col_idx in [8, 9] else body_font
                cell.border = summary_border
                
            page_sum_rows.append(current_row)
            
            # Add print page break after the page summary row if there are more records
            if idx < len(records) - 1:
                ws.row_breaks.append(Break(id=current_row))
                
            current_row += 1
            items_on_page = 0
            page_start_row = current_row
            
    # Write page summary for the last page if it was not exactly 33 items
    if items_on_page > 0:
        ws.row_dimensions[current_row].height = 20
        ws.cell(row=current_row, column=2, value="รวมแต่ละหน้า").alignment = left_align
        
        c_pamt = ws.cell(row=current_row, column=8, value=f"=SUM(H{page_start_row}:H{current_row-1})")
        c_pamt.number_format = '#,##0.00;(#,##0.00);"-";@'
        c_pamt.alignment = right_align
        c_pamt.font = summary_font
        
        c_pvat = ws.cell(row=current_row, column=9, value=f"=SUM(I{page_start_row}:I{current_row-1})")
        c_pvat.number_format = '#,##0.00;(#,##0.00);"-";@'
        c_pvat.alignment = right_align
        c_pvat.font = summary_font
        
        for col_idx in range(1, 10):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.font = summary_font if col_idx in [8, 9] else body_font
            cell.border = summary_border
            
        page_sum_rows.append(current_row)
        current_row += 1

    # Leave one blank row
    ws.row_dimensions[current_row].height = 20
    current_row += 1
    
    # Write Final "รวมทั้งหมด" Row
    ws.row_dimensions[current_row].height = 20
    ws.cell(row=current_row, column=1, value="รวมทั้งหมด").alignment = left_align
    ws.cell(row=current_row, column=1).font = summary_font
    
    # Formula for overall sum of page summaries
    sum_amt_formula = "=" + "+".join(f"H{r}" for r in page_sum_rows) if page_sum_rows else "0.00"
    c_tot_amt = ws.cell(row=current_row, column=8, value=sum_amt_formula)
    c_tot_amt.number_format = '#,##0.00;(#,##0.00);"-";@'
    c_tot_amt.alignment = right_align
    c_tot_amt.font = summary_font
    c_tot_amt.border = double_bottom_border
    
    sum_vat_formula = "=" + "+".join(f"I{r}" for r in page_sum_rows) if page_sum_rows else "0.00"
    c_tot_vat = ws.cell(row=current_row, column=9, value=sum_vat_formula)
    c_tot_vat.number_format = '#,##0.00;(#,##0.00);"-";@'
    c_tot_vat.alignment = right_align
    c_tot_vat.font = summary_font
    c_tot_vat.border = double_bottom_border
    
    # Border formatting for empty cells in totals row
    for col_idx in range(1, 8):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.border = Border(top=thin_border_side)
        
    wb.save(output_path)
